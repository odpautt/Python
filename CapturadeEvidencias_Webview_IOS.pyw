import os
from os import listdir

try:
    import openpyxl
except:
    os.system('cmd /k "echo    +--------------------+& echo    + Error libreria Openpyxl +&echo    +--------------------+&echo .&echo →Se hara upgrade de pip y se instalara Openpylx&echo →Cierre para cancelar&echo .&pause&pip install openpyxl"') 

from openpyxl import load_workbook
from openpyxl.drawing.image import Image

try:
    import easygui as es
except:
    os.system('cmd /k "echo    +--------------------+& echo    + Error libreria Easygui +&echo    +--------------------+&echo .&echo →Se hara upgrade de pip y se instalara Easygui&echo →Cierre para cancelar&echo .&pause&pip install easygui"') 


try:
    import pynput
except:
    os.system('cmd /k "echo    +--------------------+& echo    + Error libreria Pynput +&echo    +--------------------+&echo .&echo →Se hara upgrade de pip y se instalara Pynput&echo →Cierre para cancelar&echo .&pause&pip install pynput"') 

from pynput.keyboard import Key, Controller, Listener
from pynput import keyboard
try:
    import pyautogui
except:
    os.system('cmd /k "echo    +--------------------+& echo    + Error libreria Pyautogui +&echo    +--------------------+&echo .&echo →Se hara upgrade de pip y se instalara pyautogui&echo →Cierre para cancelar&echo .&pause&pip install pyautogui"') 
try:
  import cv2
except:
  os.system('cmd /k "echo    +--------------------+& echo    + Error libreria Pynput +&echo    +--------------------+&echo .&echo →Se hara upgrade de pip y se instalara OpenCV&echo →Cierre para cancelar&echo .&pause& pip install opencv-python"') 

 #********************************************************************************************************************************************************************

i = 0
path=""
name=""
caso=""
def nombre_caso():
  global name
  name=es.enterbox(msg="Digite el número del caso",title="Número de caso")
  if name== None:
      es.msgbox(msg='Ejecución finalizada no se ingreso nombre del caso',title='Toma de Evidencias', ok_button='Salir')
      exit()


def mensaje_inicial():
    es.msgbox(msg='                          *** Programa Iniciado ***\n\n\n\nOprima Impr Pant  -->  Captura Evidencia Caso Exitoso\n\nOprima Shift+f  -->  Captura Evidencia Caso Fallido\n\nOprima Esc  -->  Para finalizar el programa\n\n\n\n A contiunuación, seleccione la ruta para guardar las evidencias...',
          title='Toma de Evidencias', 
          ok_button='Continuar')

def mensaje_final():
    es.msgbox(msg='Programa Finalizado',
          title='Toma de Evidencias', 
          ok_button='Finalizar')
    #inserta_imagenes()

    exit()
def captura_dir():# Captura la ruta donde se almacenaran las imagenes
     global path
     path= es.diropenbox(msg="Abrir directorio:",
                           title="Control: diropenbox",
                           default='/home/tcs')

def captura_bug():
  global i
  global path
  global name
  screenshot = pyautogui.screenshot(region=(750,208, 430, 798))
  i =i+1
  nombre=name +"_" + str(i)+"_Fallido"
  
  punto_1=(1,1)
  punto_2=(430,798)
  color=(33,15,255)
  thickness= 10

  screenshot.save(path +"/"+ nombre + ".png")
  img=cv2.imread(path +"/"+ nombre + ".png")
  img=cv2.rectangle(img, punto_1,punto_2, color,thickness)# genera el recuadro 
  cv2.imwrite(path +"/"+ nombre + ".png",img)# guarda la imagen

def caso_nuevo():
  print("caso nuevo")
  global name, i
  es.msgbox(msg='Nuevo caso',title='Toma de Evidencias',ok_button='Continuar')
  i=0
  captura_dir()
  name=es.enterbox(msg="Digite el número del caso",title="Número de caso")


def captura_ok():
  global i
  global path
  global name
  nombre=""
  screenshot = pyautogui.screenshot(region=(750,208, 430, 798))
  i =i+1
  nombre=name +"_" + str(i)
  #puntos para generar recuadro de color verde
  punto_1=(1,1)
  punto_2=(430,798)
  color=(0,215,0)
  thickness= 10

  screenshot.save(path +"/"+ nombre + ".png")
  img=cv2.imread(path +"/"+ nombre + ".png")
  img=cv2.rectangle(img, punto_1,punto_2, color,thickness)# genera el recuadro 
  cv2.imwrite(path +"/"+ nombre + ".png",img)# guarda la imagen
  
def inserta_imagenes():

  #genera el listado de las imagenes .JPG
  listado=[f for f in listdir('.') if f.endswith('.JPG',) or f.endswith('.jpg')]
  #listado=[f for f in listado if f.find(nb)==0]#los deja las imagenes del caso de interes para agregar al archivo 
  celda=1
    
  archivo= "./ejemplo.xlsx"

  try:
    wb=load_workbook(archivo)
    #hojas=wb.get_sheet_names()
    ws=wb.active
    celda_a= ws.cell(14,1)# se declara la celda inicial para pegar la imagen "A14"


#guarda las imagenes que esten en la listas
    for i in listado:
      celda_ini=str(celda_a)
      celda_ini=celda_ini[14:16]

      img=Image(i)
      img.width=960.0
      img.height=540
      ws.add_image(img,celda_ini)
      celda_a=celda_a.offset(0,12)
      celda_ini=str(celda_ini)
      celda_ini=celda_ini[14:16]

    wb.save('ejemplo.xlsx')
    es.msgbox(msg='Evidencias Cargadas al Archivo',title='Toma de Evidencias',ok_button='Finalizar')
  except:
    print("EL archivo no se encuentra, esta abierto o se daño... por favor verifique")


mensaje_inicial()
captura_dir()
if path== None:
  es.msgbox(msg='Ejecución finalizada no se selecciono ruta de almacenamiento',
           title='Toma de Evidencias', 
           ok_button='Salir')
  exit()
nombre_caso()

hotkeys = { '<shift>+<ESC>': mensaje_final,
'<print_screen>': captura_ok ,
'<shift>+f': captura_bug}

#escuchador con la clase GlobalHotKeys
with keyboard.GlobalHotKeys(hotkeys) as l:
    l.join()