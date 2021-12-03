from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from os import listdir
import os
#listado= listdir('.')
nb="SmokeTest_LIbre destino"
listadoaux=[]
path='C:/Users/OOP07199/Documents/O.Pautt/Evidencias/OE_4636'

def inserta_imagenes():
	global listadoaux

	listado=[f for f in listdir(path) if f.endswith('.JPG',) or f.endswith('.png')]
	listadoaux= listado
	print(listado)
	# for i in range(len(listado)):
	# 	a= listado[i].find("_")
	# 	b= listado[i]
	# 	listadoaux[i]=b[0:a]
	# listadoaux=sorted(list(set(listadoaux)))#elimina los duplicados de una lista y los ordena 

	# print("listado aux: ",listadoaux)
	# nb=listadoaux[0]
	listado=[f for f in listado if f.find(nb)==0]#agregar a captura de pantallas 
	listado.sort()
	print(listado)
	#permite eliminar el .png de los nombres para poder filtarlos y dejar solo los casos de prueba


	archivo= "./ejemplo.xlsx"
	try:
		wb=load_workbook(archivo)
		#wb.active.title(nb)
		#hojas=wb.get_sheet_names()
		ws=wb.active

		celda_a= ws.cell(14,1)# se declara la celda inicial para pegar la imagen "A1"



		for i in listado:

			celda_ini=str(celda_a)
			a=celda_ini.find(".")
			b=celda_ini.find(">")
			celda_ini=celda_ini[a+1:b]
			print(celda_ini)
			img=Image(path+"/"+i)
			img.width=960.0
			img.height=540
			ws.add_image(img,celda_ini)
			celda_a=celda_a.offset(0,12)


		wb.save('ejemplo.xlsx')
	except:
		print("EL archivo no se encuentra, esta abierto o se daño... por favor verifique")
inserta_imagenes()
